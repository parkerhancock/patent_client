import datetime
import logging
import re
import sqlite3
from pathlib import Path

import lxml.etree as ET
from openpyxl import load_workbook

from patent_client import BASE_DIR
from patent_client._async.http_client import PatentClientSession
from patent_client.util.asyncio_util import run_sync

session = PatentClientSession()

legal_code_dir = BASE_DIR / "epo"
legal_code_dir.mkdir(exist_ok=True, parents=True)
db_location = legal_code_dir / "legal_codes.sqlite"


logger = logging.getLogger(__name__)


def current_date():
    return datetime.datetime.now().date()


def generate_legal_code_db():
    current = has_current_spreadsheet()
    if current:
        logger.info("Legal Code Database is Current - skipping database creation")
        return
    else:
        try:
            logger.info("Legal Code Database is out of date - creating legal code database")
            path = get_spreadsheet()
        except Exception:
            logger.exception("Could not find live code file - falling back to default dated 2023-11-05")
            path = Path(__file__).parent / "legal_code_descriptions_2023-44-0.xlsx"
        create_code_database(path)


def has_current_spreadsheet():
    con = sqlite3.connect(db_location, timeout=30)
    cur = con.cursor()
    try:
        fname = cur.execute("SELECT * FROM meta").fetchone()[0]
        date_string = re.search(r"legal_code_descriptions_([\d-]+)\.xlsx", fname).group(1)
        try:
            date = datetime.datetime.strptime(date_string, "%Y-%W-%w").date()
        except ValueError:  # Indicates using an older format and database needs to be updated
            return False
        age = datetime.datetime.now().date() - date
        logger.debug(f"Legal Code Database is {age} days old")
        return age.days < 7
    except (sqlite3.OperationalError, TypeError):
        return False


def get_spreadsheet_from_epo_website() -> tuple[datetime.date, str]:
    url = "https://www.epo.org/searching-for-patents/data/coverage/weekly.html"
    response = run_sync(session.get(url))
    response.raise_for_status()
    tree = ET.HTML(response.text)
    date_string = tree.xpath(".//tr/td[contains(text(), 'Legal event codes')]/../td[4]")[0].text.strip()
    week, year = date_string.split()[1].split("/")
    date = datetime.datetime.strptime(f"{year}-{week}-0", "%Y-%W-%w").date()
    excel_url = tree.xpath('.//a[contains(@href, "Legal-event-codes")][1]/@href')[0]
    logger.info(f"Found new spreadsheet for {date.isoformat()}: {excel_url}")
    return (date, excel_url)


def get_spreadsheet() -> tuple[datetime.date, Path]:
    date, excel_url = get_spreadsheet_from_epo_website()
    out_path = legal_code_dir / f"legal_code_descriptions_{date.strftime('%Y-%W-%w')}.xlsx"
    if out_path.exists():
        logger.info(f"File already downloaded! Current as of {date.isoformat()}")
        return out_path
    out_path = session.download(excel_url, path=out_path)
    logger.info(f"Downloaded new live code file for date {date.isoformat()}")
    return out_path


def create_code_database(excel_path):
    con = sqlite3.connect(db_location, timeout=30)
    cur = con.cursor()
    try:
        meta = cur.execute("SELECT * FROM meta").fetchone()[0]
        if meta == excel_path.name:
            logger.debug(f"Excel file {excel_path.name} already loaded. Skipping!")
            return
    except (sqlite3.OperationalError, TypeError):
        pass

    cur.execute("CREATE TABLE IF NOT EXISTS meta (file_name text)")
    cur.execute("INSERT INTO meta values (?)", (excel_path.name,))
    wb = load_workbook(excel_path)
    data = list(tuple(i.strip() for i in r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True))
    rows = data[1:]
    cur.execute(
        """CREATE TABLE IF NOT EXISTS legal_codes (
    country_code text,
    event_code text,
    date_created text,
    influence text,
    description text,
    last_update text,
    description_orig text,
    last_update_orig text,
    event_class text,
    event_class_description text)"""
    )
    cur.execute("""CREATE INDEX IF NOT EXISTS country_event_code ON legal_codes (country_code, event_code)""")
    cur.executemany("INSERT INTO legal_codes values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", rows)
    con.commit()


class LegalCodes:
    def __init__(self):
        self.connection = sqlite3.connect(db_location, timeout=30)
        self.connection.row_factory = sqlite3.Row

    def get_code_data(self, country_code, legal_code):
        cur = self.connection.cursor()
        try:
            return dict(
                cur.execute(
                    "SELECT * FROM legal_codes WHERE country_code = ? AND event_code = ?",
                    (country_code, legal_code),
                ).fetchone()
            )
        except TypeError:
            raise Exception(f"No Event Data found for {country_code} - {legal_code}")
