import datetime
import logging
import re
import sqlite3
from pathlib import Path

import lxml.etree as ET
from openpyxl import load_workbook
from patent_client import SETTINGS
from patent_client.epo.ops.session import session

dir = Path(SETTINGS.DEFAULT.BASE_DIR).expanduser() / "epo"
dir.mkdir(exist_ok=True, parents=True)
db_location = dir / "legal_codes.sqlite"


logger = logging.getLogger(__name__)


def current_date():
    return datetime.datetime.now().date()


def generate_legal_code_db():
    current = has_current_spreadsheet()
    if current:
        logger.debug("Legal Code Database is Current - skipping database creation")
        return
    else:
        logger.debug("Legal Code Database is out of date - creating legal code database")
        path = get_spreadsheet()
        create_code_database(path)


def has_current_spreadsheet():
    con = sqlite3.connect(db_location, timeout=30)
    cur = con.cursor()
    try:
        fname = cur.execute("SELECT * FROM meta").fetchone()[0]
        date_string = re.search(r"legal_code_descriptions_(\d+)\.xlsx", fname).group(1)
        date = datetime.datetime.strptime(date_string, "%Y%m%d").date()
        age = datetime.datetime.now().date() - date
        logger.debug(f"Legal Code Database is {age} days old")
        return age.days <= 30
    except (sqlite3.OperationalError, TypeError):
        return False


def get_spreadsheet():
    url = "https://www.epo.org/searching-for-patents/data/coverage/weekly.html"
    response = session.get(url)
    response.raise_for_status()
    tree = ET.HTML(response.text)
    try:
        excel_url = tree.xpath('.//*[contains(@href, "legal_code_descriptions")][1]/@href')[0]
        out_path = dir / excel_url.split("/")[-1]
        if out_path.exists():
            return out_path
        response = session.get(excel_url, stream=True)
        with out_path.open("wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        return out_path
    except IndexError:
        logger.debug("Could not find live code file - falling back to default dated 2022-11-12")
        return Path(__file__).parent / "legal_code_descriptions_20221112.xlsx"


def create_code_database(excel_path):
    con = sqlite3.connect(db_location, timeout=30)
    cur = con.cursor()
    try:
        meta = cur.execute("SELECT * FROM meta").fetchone()[0]
        if meta == excel_path.name:
            logger.debug(f"Excel file {excel_path.name} already loaded. Skipping!")
            return
    except (sqlite3.OperationalError, TypeError):
        pass

    cur.execute("CREATE TABLE IF NOT EXISTS meta (file_name text)")
    cur.execute("INSERT INTO meta values (?)", (excel_path.name,))
    wb = load_workbook(excel_path)
    data = list(tuple(i.strip() for i in r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True))
    rows = data[1:]
    cur.execute(
        """CREATE TABLE IF NOT EXISTS legal_codes (
    country_code text,
    event_code text,
    date_created text,
    influence text,
    description text,
    last_update text,
    description_orig text,
    last_update_orig text,
    event_class text,
    event_class_description text)"""
    )
    cur.execute("""CREATE INDEX IF NOT EXISTS country_event_code ON legal_codes (country_code, event_code)""")
    cur.executemany("INSERT INTO legal_codes values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", rows)
    con.commit()


class LegalCodes:
    def __init__(self):
        self.connection = sqlite3.connect(db_location, timeout=30)
        self.connection.row_factory = sqlite3.Row

    def get_code_data(self, country_code, legal_code):
        cur = self.connection.cursor()
        try:
            return dict(
                cur.execute(
                    "SELECT * FROM legal_codes WHERE country_code = ? AND event_code = ?",
                    (country_code, legal_code),
                ).fetchone()
            )
        except TypeError:
            raise Exception(f"No Event Data found for {country_code} - {legal_code}")
