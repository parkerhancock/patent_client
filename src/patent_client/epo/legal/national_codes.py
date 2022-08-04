import sqlite3
from pathlib import Path
from venv import create

import lxml.etree as ET
from openpyxl import load_workbook

from patent_client.epo.session import session

dir = Path(__file__).parent
db_location = dir / "legal_codes.sqlite"


def generate_legal_code_db():
    path = get_spreadsheet()
    create_code_database(path)


def get_spreadsheet():
    url = "https://www.epo.org/searching-for-patents/data/coverage/weekly.html"
    response = session.get(url)
    response.raise_for_status()
    tree = ET.HTML(response.text)
    excel_url = tree.xpath('.//a[contains(@href, "legal_code_descriptions")]/@href')[0]
    out_path = dir / excel_url.split("/")[-1]
    if out_path.exists():
        return out_path
    response = session.get(excel_url, stream=True)
    with out_path.open("wb") as f:
        for chunk in response.iter_content(chunk_size=8192):
            f.write(chunk)
    return out_path


def create_code_database(excel_path):
    con = sqlite3.connect(db_location)
    cur = con.cursor()
    try:
        meta = cur.execute("SELECT * FROM meta").fetchone()[0]
        if meta == excel_path.name:
            return
    except sqlite3.OperationalError:
        pass

    cur.execute("CREATE TABLE meta (file_name text)")
    cur.execute("INSERT INTO meta values (?)", (excel_path.name,))
    wb = load_workbook(excel_path)
    data = list(tuple(i.strip() for i in r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True))
    rows = data[1:]
    cur.execute("""DROP TABLE IF EXISTS legal_codes""")
    cur.execute(
        """CREATE TABLE legal_codes (
    country_code text, 
    event_code text, 
    date_created text, 
    influence text, 
    description text, 
    last_update text, 
    description_orig text, 
    last_update_orig text, 
    event_class text, 
    event_class_description text)"""
    )
    cur.execute("""CREATE INDEX country_event_code ON legal_codes (country_code, event_code)""")
    cur.executemany("INSERT INTO legal_codes values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", rows)
    con.commit()


class LegalCodes:
    def __init__(self):
        self.connection = sqlite3.connect(db_location)
        self.connection.row_factory = sqlite3.Row

    def get_code_data(self, country_code, legal_code):
        cur = self.connection.cursor()
        try:
            return dict(
                cur.execute(
                    "SELECT * FROM legal_codes WHERE country_code = ? AND event_code = ?", (country_code, legal_code)
                ).fetchone()
            )
        except TypeError:
            raise Exception(f"No Event Data found for {country_code} - {legal_code}")
